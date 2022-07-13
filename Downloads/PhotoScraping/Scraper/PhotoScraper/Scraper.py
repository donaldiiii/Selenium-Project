from openpyxl import load_workbook
from bs4 import BeautifulSoup as bs
from urllib.request import (
    urlopen, urlparse, urlunparse, urlretrieve)
import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, Text, Scrollbar, RIGHT, VERTICAL, Y, BOTH, messagebox as mb

root = tk.Tk() #whole structure of like an html doc
root.title('Photo Extractor')

fileNameArray = []
def Extractor():
    filename = filedialog.askopenfilename(initialdir="/", title="Zgjidh Excel File",
                                         filetypes=(("all files","*.xsls"), ("all files", "*.*")))
    EXCEL_PATH = filename
    excelExt = EXCEL_PATH.split(".")[-1]
    if (excelExt in ['xls','xlsx', 'xlsm']):
    #    labelLoad = tk.Label(frame, text= "Fotot Jane Duke u Ngarkuar")
    #    labelLoad.place()
    #    labelLoad.pack()
       hook(EXCEL_PATH)
    else:
       mb.showerror("Error","Sigurohuni qe tipi i file te jete Excel: '.xls' ose '.xlsx'")
    #Show Succes Message
    labelSukses = tk.Label(frame, text= "SUKSES")
    label = tk.Label(frame, text= str(len(fileNameArray)) + " FOTO U EKSTRAKTUAN ME SUKSES")
    labelSukses.pack()
    label.pack()
canvas = tk.Canvas(root, height=250, width=400, bg="#263D42", scrollregion=(0,0,1000,1000)) # create canvas

frame = tk.Frame(root, bg="white")
frame.place(relwidth=0.8, relheight =0.8, relx=0.1, rely=0.1)
canvas.pack(expand=True) # attach it to the root

openFile = tk.Button(root, text="Ngarko Excel File", padx = 5, #krijimi i butonit
                     pady = 5, fg="white",bg="#263D42", command=Extractor)
openFile.pack() #attach it

#Domain Object, properties that come from Excel
class Data(object):
    def __init__(self, Kodi=None, URL_produit=None, Reference=None):
        self.Kodi = Kodi
        self.URL_produit = URL_produit
        self.Reference = Reference



#Creates an array of all the needed Rows from Excel
def GetRecordsFromExcel(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH.strip())
        sheet = wb.get_sheet_by_name('Sheet1')
        ws = sheet
        header = [cell.value for cell in ws[1]]
        datarows = []
        for row in list(ws.rows)[1:]:
            values = {}
            for key, cell in zip(header, row):
                if( key in ['Kodi','URL_produit','Reference']):
                    values[key] = cell.value
            datarow = Data(**values)
            if not (datarow.Kodi == None and datarow.Reference == None and datarow.URL_produit == None):
                datarows.append(datarow)
        if not len(datarows) == 0:
            return datarows
        else:
            return mb.showerror("Error", "Ju lutem kontrolloni formatin e kolonave")

### Per cdo rekord, nga Url ekstraktohet foto perkatese dhe hidhet ne folderin ImagesFolder
def main(url, out_folder="ImagesFolder"):

    soup = bs(urlopen(url))
    parsed = list(urlparse(url))
    # for image in soup.findAll("img",{'class': 'slider-main-img'}):
    for image in soup.findAll("img",{'data-slidemain-index': '0'}):
        print("Image: %(src)s" % image)
        filename = soup.find("h1",{'class': 'item-title'}).text
        ext =  "." + image["src"].split(".")[-1]
        filename = re.sub("\n|\r|\t", " ", filename).strip().replace("  "," ")  + ext
        if filename in fileNameArray:
            filename = filename.replace(ext,"") + str(len(fileNameArray)) + ext
        fileNameArray.append(filename)
        parsed[2] = image["src"]
        outpath = os.path.join(out_folder, filename)
        if image["src"].lower().startswith("http"):
            urlretrieve(image["src"], outpath)
        else:
            urlretrieve(urlunparse(parsed), outpath)

def _usage():
     print("py Scraper.py")

#Init Kalon ne metoden kryesore main nje nga nje rekordet per tu procesuar
def hook(EXCEL_PATH):
    # if __name__ == "__main__":
    
    dataRows = GetRecordsFromExcel(EXCEL_PATH)
    out_folder = "ImagesFolder"
    if not os.path.exists(out_folder):
            os.makedirs(out_folder)
    for data in dataRows:
            if data.URL_produit.lower().startswith("http"):
                main(data.URL_produit,out_folder)
            else:
                mb.showerror("Error", "Sigurohu qe adresa Url te jete ne formatin e duhur (te filloje me http)")
                break
labelSukses = tk.Label(frame, text= "Ngarkoni Excel-in dhe prisni sa te ekstraktohen Fotot")
labelSukses.pack()
root.mainloop() #run GUI

